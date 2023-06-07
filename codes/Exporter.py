from genericpath import exists
import pandas as pd
import openpyxl
import os.path

#Object that contains column title, array of plant measurement data and column number
class PlantData:
    def __init__(self,title,startCol, data):      
        if startCol < 2:
            raise Exception("Data Column must be 2 or higher")
        self.title = title
        self.startCol = startCol
        self.data = data


def printPlantData(plantArray, filePathAndName, startRow, saveToExistingFile):
    #data needs to start at row 2 or higher due to titles
    if startRow < 2:
        raise Exception("Data must start on row 2 or higher")

    
    if saveToExistingFile and os.path.exists(filePathAndName):
        wb = openpyxl.load_workbook(filePathAndName);
    else:
        wb = openpyxl.Workbook()

    sheet1 = wb.active

    #get longest data array for knowing where to write the averages 
    longest = getLongestDataArrayLength(plantArray)
    avgLine = startRow + longest + 4
    firstCol = getFirstColumn(plantArray)
    
    #write Average text first 
    sheet1.cell(row =avgLine, column= firstCol-1, value="Average")

    for p in plantArray:
        rownum = startRow
        #print column title
        sheet1.cell(row=rownum,column=p.startCol, value=p.title)
        rownum+=1
        
        for d in p.data:
            #print data 
            sheet1.cell(row=rownum,column=p.startCol,value=d)
            rownum+=1

        #print average 
        sheet1.cell(row= avgLine, column= p.startCol, value = getAverage(p.data))

    #save the sheet     
    wb.save(filePathAndName)

def getLongestDataArrayLength(plantArray):
    longest = 0
    #find array with longest length
    for p in plantArray:
        if len(p.data) > longest:
            longest = len(p.data)
    
    return longest

def getFirstColumn(plantArray):
    firstCol = 9999
    #find first column in case they aren't in order 
    for p in plantArray:
        if p.startCol < firstCol:
            firstCol = p.startCol

    return firstCol


def getAverage(data):
    return sum(data) / len(data)



#test data 
data1 = [4.559, 2.562, 3.283]
pd1 = PlantData("Col-0",5,data1)

data2 = [2.551, 4.652, 1.283, 4.63, 5.72]
pd2 = PlantData("1 µM",6,data2)



#end of test data 

file = r'C:\Users\jradi\Documents\CITS5551\Exp16.xlsx'
plantArray = [pd1,pd2]

printPlantData(plantArray,file,3,True)

